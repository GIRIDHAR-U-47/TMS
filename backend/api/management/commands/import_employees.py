import openpyxl
from django.core.management.base import BaseCommand
from api.models import Employee
from datetime import datetime

class Command(BaseCommand):
    help = 'Import employees from Excel file'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str)

    def handle(self, *args, **kwargs):
        excel_file = kwargs['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        column_map = {
            'EMP NO': 'emp_no',
            'NAME': 'name',
            'GENDER': 'gender',
            'DOB': 'dob',
            'AGE': 'age',
            'DOJ': 'doj',
            'DOL': 'dol',
            'PLANT': 'plant',
            'AREA OF WORK': 'area_of_work',
            'DEPT': 'dept',
            'CATEGORY': 'category',
            'BATCH NO': 'batch_no',
            'NO OF DAYS TRAINING': 'training_days',
            'SL 1 MARK': 'sl1_marks',
            'SL2 MARK': 'sl2_marks',
            'SL 2 OJT': 'sl2_ojt',
            'AFTER OJT AREA OF WORK': 'after_ojt_area_of_work',
            'Over all %': 'overall_percent',
            'SKILL LEVEL': 'skill_level',
            # Add more mappings as needed
        }
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_data = dict(zip(header, row))
            data = {column_map.get(k, k): v for k, v in raw_data.items() if column_map.get(k, k) in Employee._meta.fields}
            # Convert date fields
            for date_field in ['dob', 'doj', 'dol']:
                if data.get(date_field) and not isinstance(data[date_field], datetime):
                    try:
                        data[date_field] = datetime.strptime(str(data[date_field]), '%Y-%m-%d').date()
                    except Exception:
                        data[date_field] = None
            # Convert numeric fields
            for int_field in ['age', 'training_days', 'sl1_marks', 'sl2_marks']:
                if data.get(int_field) is not None:
                    try:
                        data[int_field] = int(data[int_field])
                    except Exception:
                        data[int_field] = 0
            if data.get('overall_percent') is not None:
                try:
                    data['overall_percent'] = float(data['overall_percent'])
                except Exception:
                    data['overall_percent'] = 0.0
            emp_no = data.get('emp_no')
            if not emp_no:
                continue
            Employee.objects.update_or_create(
                emp_no=emp_no,
                defaults=data
            )
        self.stdout.write(self.style.SUCCESS('Employee import complete!'))