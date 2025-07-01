from django.contrib import admin, messages
from django.urls import path
from django.shortcuts import render, redirect
from .models import (
    Employee, TrainingModule, EmployeeTrainingModule, TrainingRecord,
    OnJobTraining, DexterityAssessment, PerformanceRecord
)
import openpyxl
from datetime import datetime


@admin.register(Employee)
class EmployeeAdmin(admin.ModelAdmin):
    list_display = ['emp_no', 'name', 'area_of_work', 'plant', 'category', 'skill_level', 'overall_percent', 'created_at']
    list_filter = ['area_of_work', 'plant', 'category', 'skill_level', 'gender', 'created_at']
    search_fields = ['emp_no', 'name', 'batch_no']
    readonly_fields = ['created_at', 'updated_at']
    fieldsets = (
        ('Basic Information', {
            'fields': ('emp_no', 'name', 'gender', 'dob', 'age', 'doj', 'dol', 'photo')
        }),
        ('Work Details', {
            'fields': ('plant', 'area_of_work', 'category', 'batch_no')
        }),
        ('Training Details', {
            'fields': ('training_days', 'sl1_marks', 'sl2_marks', 'sl2_ojt', 'after_ojt_area_of_work', 'overall_percent', 'skill_level')
        }),
        ('SL Assessment', {
            'fields': ('sl1_status', 'sl2_status', 'sl3_status')
        }),
        ('Additional Information', {
            'fields': ('remarks', 'created_at', 'updated_at')
        }),
    )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('upload-excel/', self.admin_site.admin_view(self.upload_excel))
        ]
        return custom_urls + urls

    def upload_excel(self, request):
        if request.method == "POST" and request.FILES.get("excel_file"):
            try:
                wb = openpyxl.load_workbook(request.FILES["excel_file"])
                ws = wb.active
                header = [str(cell.value).strip().upper() if cell.value else '' for cell in ws[1]]
                
                column_map = {
                    'EMP NO': 'emp_no',
                    'NAME': 'name',
                    'GENDER': 'gender',
                    'DOB': 'dob',
                    'AGE': 'age',
                    'DOJ': 'doj',
                    'DOL': 'dol',
                    'PLANT': 'plant',
                    'AREA OF WORK': 'area_of_work',
                    'DEPT': 'category',
                    'CATEGORY': 'category',
                    'BATCH NO': 'batch_no',
                    'NO OF DAYS TRAINING': 'training_days',
                    'SL 1 MARK': 'sl1_marks',
                    'SL2 MARK': 'sl2_marks',
                    'SL 2 OJT': 'sl2_ojt',
                    'AFTER OJT DEPT': 'after_ojt_area_of_work',
                    'OVER ALL %': 'overall_percent',
                    'SKILL LEVEL': 'skill_level',
                    'SL1 STATUS': 'sl1_status',
                    'SL2 STATUS': 'sl2_status',
                    'SL3 STATUS': 'sl3_status',
                    'REMARKS': 'remarks'
                }

                success_count = 0
                error_count = 0
                error_messages = []

                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        raw_data = dict(zip(header, row))
                        data = {column_map.get(k.strip().upper(), k): v 
                               for k, v in raw_data.items() 
                               if column_map.get(k.strip().upper()) in Employee._meta.fields}
                        
                        # Skip empty rows
                        if not data.get('emp_no'):
                            continue

                        # Convert date fields
                        for date_field in ['dob', 'doj', 'dol']:
                            if data.get(date_field):
                                try:
                                    if isinstance(data[date_field], datetime):
                                        data[date_field] = data[date_field].date()
                                    else:
                                        data[date_field] = datetime.strptime(str(data[date_field]), '%Y-%m-%d').date()
                                except ValueError:
                                    data[date_field] = None
                                    error_messages.append(f"Row {row_idx}: Invalid date format for {date_field}")

                        # Convert numeric fields
                        for int_field in ['age', 'training_days', 'sl1_marks', 'sl2_marks']:
                            if data.get(int_field) is not None:
                                try:
                                    data[int_field] = int(float(str(data[int_field])))
                                except (ValueError, TypeError):
                                    data[int_field] = 0
                                    error_messages.append(f"Row {row_idx}: Invalid numeric value for {int_field}")

                        # Convert overall_percent to float
                        if data.get('overall_percent') is not None:
                            try:
                                data['overall_percent'] = float(str(data['overall_percent']).rstrip('%'))
                            except (ValueError, TypeError):
                                data['overall_percent'] = 0.0
                                error_messages.append(f"Row {row_idx}: Invalid percentage value for overall_percent")

                        # Handle empty strings
                        for field in ['sl2_ojt', 'after_ojt_area_of_work', 'remarks', 'batch_no']:
                            if field in data and (data[field] is None or data[field] == ''):
                                data[field] = None

                        emp_no = data.get('emp_no')
                        if emp_no:
                            Employee.objects.update_or_create(
                                emp_no=emp_no,
                                defaults=data
                            )
                            success_count += 1
                    except Exception as e:
                        error_count += 1
                        error_messages.append(f"Row {row_idx}: {str(e)}")

                if error_messages:
                    self.message_user(
                        request, 
                        f"Excel import completed with {success_count} successes and {error_count} errors. Errors: {'; '.join(error_messages)}",
                        messages.WARNING
                    )
                else:
                    self.message_user(
                        request,
                        f"Excel import completed successfully! {success_count} records processed.",
                        messages.SUCCESS
                    )
            except Exception as e:
                self.message_user(
                    request,
                    f"Error processing Excel file: {str(e)}",
                    messages.ERROR
                )
            return redirect("..")
        return render(request, "admin/excel_upload.html")


@admin.register(TrainingModule)
class TrainingModuleAdmin(admin.ModelAdmin):
    list_display = ['s_no', 'title', 'expert']
    list_filter = ['expert']
    search_fields = ['title', 'expert']
    ordering = ['s_no']


@admin.register(EmployeeTrainingModule)
class EmployeeTrainingModuleAdmin(admin.ModelAdmin):
    list_display = ['employee', 'module', 'status', 'completed_date']
    list_filter = ['status', 'completed_date', 'module']
    search_fields = ['employee__name', 'employee__emp_no', 'module__title']


@admin.register(TrainingRecord)
class TrainingRecordAdmin(admin.ModelAdmin):
    list_display = ['employee', 'date', 'training_program', 'duration']
    list_filter = ['date', 'employee__area_of_work']
    search_fields = ['employee__name', 'training_program']


@admin.register(OnJobTraining)
class OnJobTrainingAdmin(admin.ModelAdmin):
    list_display = ['employee', 'product_process', 'machine_operations', 'quality_check_points']
    list_filter = ['employee__area_of_work', 'employee__plant']
    search_fields = ['employee__name', 'product_process', 'machine_operations']


@admin.register(DexterityAssessment)
class DexterityAssessmentAdmin(admin.ModelAdmin):
    list_display = ['employee', 'basic_skills_total', 'advanced_skills_total', 'overall_score', 'created_at']
    list_filter = ['created_at', 'employee__area_of_work']
    search_fields = ['employee__name', 'employee__emp_no']
    readonly_fields = ['basic_skills_total', 'advanced_skills_total', 'overall_score', 'created_at', 'updated_at']
    fieldsets = (
        ('Employee', {
            'fields': ('employee',)
        }),
        ('Basic Skills', {
            'fields': ('test_1s_2s', 'test_1s_2s_ball', 'memory_test', 'mind_hand_coordination', 
                      'nerve_stability', 'material_identification', 'pick_place_sequence', 
                      'pick_right_material', 'visual_inspection', 'defect_identification', 'written_test')
        }),
        ('Advanced Skills', {
            'fields': ('insert_loading_1', 'insert_loading_2', 'safety_test', 'painting',
                      'screw_assembly', 'air_cleaner_assembly', 'msa_test', 'deflashing')
        }),
        ('Totals', {
            'fields': ('basic_skills_total', 'advanced_skills_total', 'overall_score')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at')
        }),
    )


@admin.register(PerformanceRecord)
class PerformanceRecordAdmin(admin.ModelAdmin):
    list_display = ['employee', 'day', 'operation_name', 'final_score', 'supervisor_approved', 'personnel_certified']
    list_filter = ['day', 'supervisor_approved', 'personnel_certified', 'employee__area_of_work']
    search_fields = ['employee__name', 'operation_name', 'description']
    readonly_fields = ['created_at', 'updated_at']
    fieldsets = (
        ('Employee Information', {
            'fields': ('employee', 'day')
        }),
        ('Performance Details', {
            'fields': ('description', 'su_status', 'scope', 'operation_name', 'production', 
                      'weight', 'quantity', 'pro_n', 'perf_n', 'final_score')
        }),
        ('Approval', {
            'fields': ('supervisor_approved', 'personnel_certified')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at')
        }),
    )
